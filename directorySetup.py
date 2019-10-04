import os
import openpyxl
import shelve


def create_files(datad, hidden):
    os.makedirs(datad, mode=hidden, exist_ok=True)
    print('new directory added '+datad)
    if datad == 'C:\\AnniqueSpreadSheets\\data':
        os.system("attrib +h C:\\AnniqueSpreadSheets\\data")

def create_workbook(getd):
    os.chdir(getd)
    workbook = openpyxl.Workbook()
    workbook.create_sheet('Sheet1')
    workbook.__delitem__('Sheet')
    print(workbook.sheetnames)
    workbook.save('AnniqueTargetRun.xlsx')


def create_shelve(datad, sname):
    os.chdir(datad)
    createShelve = shelve.open(sname)
    createShelve.__setitem__('month6', [])
    createShelve.__setitem__('month5', [])
    createShelve.__setitem__('month4', [])
    createShelve.__setitem__('month3', [])
    createShelve.__setitem__('month2', [])
    createShelve.__setitem__('month1', [])
    createShelve.__setitem__('date', 0)

    createShelve.__setitem__('cmonth6', None)
    createShelve.__setitem__('cmonth5', None)
    createShelve.__setitem__('cmonth4', None)
    createShelve.__setitem__('cmonth3', None)
    createShelve.__setitem__('cmonth2', None)
    createShelve.__setitem__('cmonth1', None)
    createShelve.close()


def readme(directory):
    text = open(directory+'\\ReadMe.txt', 'w')
    text.write('This software was designed to be used with Annique\'s data and will most likely not be usable in other '
               'contexts.\n\n '
               'In order to use this software please look at the following list\n'
               '1. An excel spread sheet called AnniqueTargetRun has been created under {0}\\get. use\n this to store'
               ' all information you wish to be used\n'
               '2. Each time you run this programme a new spread sheet will be created under {0}\\set\n it will be'
               ' named today\'s date \nNOTE:\"only one can exist per day as 2 files may not have the same '
               'name. Running the programme twice will result in the first file being overwritten\"\n'
               '3. Every time this programme is run in a new month the previous month\'s sales will come up as a '
               'new column'
               '4. Do not modify anything in the data file as this is used for setting up parts of the programme and '
               'may not function properly if they are tampered with.\n'
               '5. If all else fails deleting {0} and re running the programme will reset everything\n\n'
               'I hope all of this helps ;)\n\n\n\n\n\n\n'
               'software developed by Leonard Brugman'.format(directory))
