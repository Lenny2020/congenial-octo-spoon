import openpyxl
import os
import re
import datetime
import shelve
import directorySetup
from os import path
from openpyxl.styles import PatternFill


def create_spreadsheet(column, rows):
    for j in range(len(column)):
        sheetFinal.cell(row=1, column=j+1).value = column[j]

        for k in range(len(rows)):

            sheetFinal.cell(row=k+3, column=j+1).value = rows[k][j+1]
            if j == 3:

                # sets cell colour to red if target not reached
                if rows[k][0]:

                    sheetFinal.cell(row=k+3, column=j+1).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE',
                                                                            fill_type="solid")
                if rows[k][0] is False:
                    sheetFinal.cell(row=k+3, column=j+1).fill = PatternFill(start_color='ffffff', end_color='ffffff',
                                                                            fill_type="solid")


def remove_decimals(word):
    periodRegex = re.compile(r'[.]')
    commaRegex = re.compile(r',')
    word = commaRegex.sub('', word)
    word = periodRegex.sub('', word)
    word = int(word)
    return word


date = datetime.datetime.now()
# date = datetime.datetime(2019, 9, 12)
directory = 'C:\\AnniqueSpreadSheets'
getDirectory = directory+'\\Get'
setDirectory = directory+'\\Set'
dataDirectory = directory+'\\data'

if path.exists(directory) is False:
    directorySetup.create_files(dataDirectory, 0o77)
    directorySetup.create_files(getDirectory, 0o77)
    directorySetup.create_files(setDirectory, 0o77)
    directorySetup.create_workbook(getDirectory)
    directorySetup.readme(directory)
    directorySetup.create_shelve(dataDirectory, 'income')


os.chdir(getDirectory)
getSheet = 'AnniqueTargetRun.xlsx'
wb = openpyxl.load_workbook(getSheet)
sheet = wb['Sheet1']

os.chdir(dataDirectory)
columnShelf = shelve.open('income')

cmonth6 = columnShelf.__getitem__('cmonth6')
cmonth5 = columnShelf.__getitem__('cmonth5')
cmonth4 = columnShelf.__getitem__('cmonth4')
cmonth3 = columnShelf.__getitem__('cmonth3')
cmonth2 = columnShelf.__getitem__('cmonth2')
cmonth1 = columnShelf.__getitem__('cmonth1')

month6 = columnShelf.__getitem__('month6')
month5 = columnShelf.__getitem__('month5')
month4 = columnShelf.__getitem__('month4')
month3 = columnShelf.__getitem__('month3')
month2 = columnShelf.__getitem__('month2')
month1 = columnShelf.__getitem__('month1')

if date.month != columnShelf.__getitem__('date'):

    columnShelf.__setitem__('month6', month5)
    columnShelf.__setitem__('month5', month4)
    columnShelf.__setitem__('month4', month3)
    columnShelf.__setitem__('month3', month2)
    columnShelf.__setitem__('month2', month1)

    strDay = str(date.year) + '-' + str(date.month)
    columnShelf.__setitem__('cmonth6', cmonth5)
    columnShelf.__setitem__('cmonth5', cmonth4)
    columnShelf.__setitem__('cmonth4', cmonth3)
    columnShelf.__setitem__('cmonth3', cmonth2)
    columnShelf.__setitem__('cmonth2', cmonth1)
    columnShelf.__setitem__('cmonth1', strDay)

    columnShelf.__setitem__('date', date.month)

    print('new date set')


cmonth6 = columnShelf.__getitem__('cmonth6')
cmonth5 = columnShelf.__getitem__('cmonth5')
cmonth4 = columnShelf.__getitem__('cmonth4')
cmonth3 = columnShelf.__getitem__('cmonth3')
cmonth2 = columnShelf.__getitem__('cmonth2')
cmonth1 = columnShelf.__getitem__('cmonth1')

month6 = columnShelf.__getitem__('month6')
month5 = columnShelf.__getitem__('month5')
month4 = columnShelf.__getitem__('month4')
month3 = columnShelf.__getitem__('month3')
month2 = columnShelf.__getitem__('month2')
month1 = columnShelf.__getitem__('month1')

os.chdir(setDirectory)
saveSheet = str(date.date())+'.xlsx'
wbwrite = openpyxl.Workbook()
wbwrite.save(saveSheet)
wbwrite = openpyxl.load_workbook(saveSheet)
sheetFinal = wbwrite['Sheet']


# sets the columns
columns = ['MemberNo', 'Name', 'level', 'Career Sales Sales', 'Personal Sales', 'Target Reached',
           'Added Discount', 'Next Target', 'Amount to GO']
if cmonth6:
    columns.append(cmonth6)

if cmonth5:
    columns.append(cmonth5)

if cmonth4:
    columns.append(cmonth4)

if cmonth3:
    columns.append(cmonth3)

if cmonth2:
    columns.append(cmonth2)

data = []

# sets string to int

numRegex = re.compile(r'\d+')

# loops through all the rows
for i in range(3, sheet.max_row):
    person = []
    discountOff = 0
    step = 0
    nextTarget = 3301
    reachedTarget = 0
    colour = False
    # gets all the info from original table
    memberNum = sheet.cell(row=i, column=1).value
    name = sheet.cell(row=i, column=2).value
    level = str(sheet.cell(row=i, column=3).value)
    career = str(sheet.cell(row=i, column=7).value)
    income = str(sheet.cell(row=i, column=12).value)

    # manipulates data into required fields avoiding any none data type
    if name:
        try:
            level = numRegex.search(level)
            level = int(level.group())
            career = remove_decimals(career)
            income = remove_decimals(income)
        except ValueError:
            a = 0

        if income <= 3300:
            discountOff = 0
            reachedTarget = 0
            step = nextTarget - income

        elif income <= 6999:
            discountOff = 5
            reachedTarget = 3301
            nextTarget = 7000
            step = nextTarget - income
        elif income <= 13999:
            discountOff = 10
            reachedTarget = 7000
            nextTarget = 14000
            step = nextTarget - income
        elif income <= 26999:
            discountOff = 13
            reachedTarget = 14000
            nextTarget = 27000
            step = nextTarget - income
        elif income <= 39999:
            discountOff = 16
            reachedTarget = 27000
            nextTarget = 40000
            step = nextTarget - income
        else:
            discountOff = 20
            reachedTarget = 40000

        if career < 3000:
            colour = True

        # stores all variables in a list to be grouped together
        person = [colour, memberNum, name, level, career, income, reachedTarget,
                  discountOff, nextTarget, step]
        try:
            if month6:
                person.append(month6[i-3])

            if month5:
                person.append(month5[i-3])

            if month4:
                person.append(month4[i-3])

            if month3:
                person.append(month3[i-3])

            if month2:
                person.append(month2[i-3])
        except IndexError:
            person.append(0)

        # stores list in list
        data.append(person)

# sorts the list alphabetically

create_spreadsheet(columns, data)


storeIncome = []
for i in data:
    storeIncome.append(i[5])


os.chdir(dataDirectory)


month5 = columnShelf.__getitem__('month5')
month4 = columnShelf.__getitem__('month4')
month3 = columnShelf.__getitem__('month3')
month2 = columnShelf.__getitem__('month2')
month1 = columnShelf.__getitem__('month1')

cmonth5 = columnShelf.__getitem__('cmonth5')
cmonth4 = columnShelf.__getitem__('cmonth4')
cmonth3 = columnShelf.__getitem__('cmonth3')
cmonth2 = columnShelf.__getitem__('cmonth2')
cmonth1 = columnShelf.__getitem__('cmonth1')


columnShelf.__setitem__('month1', storeIncome)


columnShelf.close()
os.chdir(setDirectory)
wbwrite.save(saveSheet)
